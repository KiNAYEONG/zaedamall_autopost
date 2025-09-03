# tools/mall_auto_write.py
# -*- coding: utf-8 -*-
"""
재다몰 자동 업로드 (견고 버전)
- .env의 CHROME_USER_DATA_DIR + CHROME_PROFILE로 1차 실행
  -> "in use" 또는 crash 시 CHROME_FALLBACK_DIR로 폴백
- 미로그인/권한 알럿 자동 처리
- 리스트에서 '글쓰기' 버튼 클릭 방식 + write.php 직접 진입 방식 모두 지원
- docs/data.xlsx에서 A(제목)/B(본문) 읽고 C 상태가 DONE/PUBLISHED/SKIP가 아닌 첫 행을 업로드
"""

from __future__ import annotations
from dotenv import load_dotenv
import os, sys, time, argparse, datetime
from pathlib import Path
from urllib.parse import urlparse, parse_qs
import openpyxl

from selenium.webdriver import Chrome, ChromeOptions
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    SessionNotCreatedException,
    WebDriverException,
    NoSuchElementException,
    TimeoutException,
    UnexpectedAlertPresentException,
)

from webdriver_manager.chrome import ChromeDriverManager

# ──────────────────────────────
# 기본 경로/상수
# ──────────────────────────────
ROOT = Path(__file__).resolve().parent.parent
DOCS = ROOT / "docs"
XLSX = DOCS / "data.xlsx"
MAX_WAIT = 20


def log(msg: str):
    print(msg, flush=True)


# ──────────────────────────────
# 엑셀 헬퍼
# ──────────────────────────────
def load_next_row():
    if not XLSX.exists():
        raise FileNotFoundError(f"엑셀 파일이 없습니다: {XLSX}")
    wb = openpyxl.load_workbook(XLSX)
    ws = wb.active
    for i in range(2, ws.max_row + 1):
        title = (ws[f"A{i}"].value or "").strip()
        body = (ws[f"B{i}"].value or "").strip()
        status = (ws[f"C{i}"].value or "").strip().upper()
        if title and body and status not in ("DONE", "PUBLISHED", "SKIP"):
            return wb, ws, i, title, body
    return wb, ws, None, None, None


def mark_done(wb, ws, row: int):
    ws[f"C{row}"] = "DONE"
    ws[f"D{row}"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    wb.save(XLSX)


# ──────────────────────────────
# 공용 Selenium 헬퍼
# ──────────────────────────────
def wait_ready(drv, timeout: int = MAX_WAIT):
    WebDriverWait(drv, timeout).until(
        lambda d: d.execute_script("return document.readyState") == "complete"
    )


def accept_all_alerts(drv, max_loops: int = 5):
    """열려있는 JS alert/confirm이 있으면 전부 수락."""
    for _ in range(max_loops):
        try:
            a = drv.switch_to.alert
            txt = a.text
            try:
                a.accept()
            except Exception:
                pass
            log(f"⚠ 알럿 감지 → 자동 수락: {txt}")
            time.sleep(0.6)
        except Exception:
            break


def safe_get(drv, url: str, timeout: int = MAX_WAIT):
    drv.get(url)
    try:
        wait_ready(drv, timeout)
    finally:
        accept_all_alerts(drv)


def find_first(drv, selectors: list[str], by: By = By.CSS_SELECTOR, wait_s: int = 8):
    """selectors를 순회하며 첫 번째로 존재하는 요소를 반환."""
    for sel in selectors:
        try:
            el = WebDriverWait(drv, wait_s).until(
                EC.presence_of_element_located((by, sel))
            )
            return el, sel
        except TimeoutException:
            continue
    raise NoSuchElementException(f"해당 셀렉터들을 찾을 수 없습니다: {selectors}")


# ──────────────────────────────
# 로그인 감지/시도
# ──────────────────────────────
def is_logged_in(drv) -> bool:
    """상단 네비/페이지 어디서든 '로그아웃' 또는 logout 링크가 보이면 로그인 상태로 간주."""
    try:
        # 빠른 텍스트 검사 (헤더/푸터 포함)
        html = drv.page_source
        if "로그아웃" in html or "logout" in html.lower():
            return True
        # 링크 형태
        links = drv.find_elements(By.XPATH, "//a[contains(@href,'logout') or contains(.,'로그아웃')]")
        return len(links) > 0
    except Exception:
        return False


def try_auto_login(drv, home_url: str = "https://zae-da.com/") -> bool:
    """환경변수 MALL_ID/MALL_PW를 사용해 자동 로그인 시도."""
    uid = os.getenv("MALL_ID", "").strip()
    pw = os.getenv("MALL_PW", "").strip()
    if not uid or not pw:
        return False

    # 홈 → '로그인' 클릭 (없으면 바로 로그인 폼으로 진입)
    safe_get(drv, home_url)
    time.sleep(0.8)

    # 로그인 링크 찾아보기
    try:
        login_link, _ = find_first(
            drv,
            [
                "//a[contains(.,'로그인')]",
                "//a[contains(@href,'login') or contains(@href,'member/login')]",
                "//button[contains(.,'로그인')]",
            ],
            by=By.XPATH,
            wait_s=5,
        )
        login_link.click()
        time.sleep(0.8)
    except Exception:
        # 링크 못 찾으면 혹시 이미 로그인 폼일 수 있으니 그대로 진행
        pass

    wait_ready(drv)

    # 아이디/비번 입력 필드 탐색
    id_sels = [
        "input[name='mb_id']",
        "input#mb_id",
        "input[name='login_id']",
        "input#login_id",
        "input[name='user_id']",
        "input[name='id']",
    ]
    pw_sels = [
        "input[name='mb_password']",
        "input#mb_password",
        "input[name='login_pw']",
        "input#login_pw",
        "input[name='user_pw']",
        "input[name='password']",
        "input[name='passwd']",
    ]
    btn_sels = [
        "//button[contains(.,'로그인')]",
        "//input[@type='submit' and (contains(@value,'로그인') or contains(@value,'login'))]",
        "//a[contains(@onclick,'login') and contains(.,'로그인')]",
    ]

    try:
        id_el, _ = find_first(drv, id_sels, By.CSS_SELECTOR, wait_s=6)
        pw_el, _ = find_first(drv, pw_sels, By.CSS_SELECTOR, wait_s=6)
        id_el.clear(); id_el.send_keys(uid)
        pw_el.clear(); pw_el.send_keys(pw)

        try:
            btn, _ = find_first(drv, btn_sels, By.XPATH, wait_s=4)
            btn.click()
        except Exception:
            # 엔터로 제출
            pw_el.submit()

        # 로그인 결과 대기 (최대 20초)
        for _ in range(20):
            time.sleep(1.0)
            accept_all_alerts(drv)
            if is_logged_in(drv):
                log("🔐 자동 로그인 성공")
                return True
        return False
    except Exception:
        return False


def wait_until_logged_in(drv, timeout_s: int = 180) -> bool:
    """수동 로그인(다른 창/현재 창) 완료를 텍스트로 감지. 키보드 입력 없이 폴링."""
    log("⏳ 로그인 감지 대기 중... (최대 3분)")
    t0 = time.time()
    while time.time() - t0 < timeout_s:
        time.sleep(2.0)
        accept_all_alerts(drv)
        try:
            if is_logged_in(drv):
                log("🔓 로그인 감지됨")
                return True
        except Exception:
            pass
    return False


def ensure_login(drv, list_url: str, write_url: str):
    """로그인 필요 시 자동 로그인 시도 → 실패하면 수동 로그인 감지."""
    # 1) 현재 로그인 상태면 바로 리턴
    try:
        accept_all_alerts(drv)
        if is_logged_in(drv):
            return
    except Exception:
        pass

    # 2) 자동 로그인 시도
    if try_auto_login(drv):
        return

    # 3) 자동 실패 시: 리스트 페이지 오픈 후 '로그인' 유도, 수동 로그인 감지
    safe_get(drv, list_url or "https://zae-da.com/")
    log("👉 로그인 페이지로 이동해 수동 로그인 해주세요. (최대 3분 내 자동 감지)")
    if not wait_until_logged_in(drv, timeout_s=180):
        raise RuntimeError("로그인을 감지하지 못했습니다. 로그인 후 다시 실행해주세요.")


# ──────────────────────────────
# 글쓰기 페이지 진입
# ──────────────────────────────
def board_id_from_url(url: str) -> str | None:
    try:
        q = parse_qs(urlparse(url).query)
        bid = q.get("boardid", [None])[0]
        return bid
    except Exception:
        return None


def goto_write_from_list(drv, list_url: str, boardid: str | None) -> bool:
    safe_get(drv, list_url)
    # 리스트에서 '글쓰기' 버튼 또는 write.php 링크 찾기
    candidates = [
        # 직접 링크
        f"//a[contains(@href,'board_write.php') and contains(@href,'boardid={boardid}')]" if boardid else "",
        # 텍스트/아이콘 버튼
        "//a[contains(.,'글쓰기') or contains(.,'작성')]",
        "//button[contains(.,'글쓰기') or contains(.,'작성')]",
        "//a[@class='btn' and (contains(.,'글쓰기') or contains(.,'작성'))]",
        "//a[contains(@class,'write')]",
    ]
    candidates = [c for c in candidates if c]

    for xpath in candidates:
        try:
            btn = WebDriverWait(drv, 8).until(EC.element_to_be_clickable((By.XPATH, xpath)))
            btn.click()
            time.sleep(0.8)
            wait_ready(drv)
            accept_all_alerts(drv)
            # write 페이지 판단: URL 또는 제목 필드 존재
            if "board_write.php" in drv.current_url:
                return True
            # 필드 존재 확인
            _ = find_first(drv, ["input[name='wr_subject']", "input[name='subject']", "input[name='title']"], By.CSS_SELECTOR, 3)
            return True
        except UnexpectedAlertPresentException:
            accept_all_alerts(drv)
            # 권한 알럿이면 로그인 시도 후 재시도
            ensure_login(drv, list_url, "")
            return goto_write_from_list(drv, list_url, boardid)
        except Exception:
            continue
    return False


def ensure_write_page(drv, list_url: str, write_url: str):
    """리스트→버튼 클릭 우선, 실패 시 write.php 직접 진입."""
    bid = board_id_from_url(write_url)
    # A) 이미 write 페이지면 통과
    try:
        if "board_write.php" in drv.current_url:
            return
    except Exception:
        pass

    # B) 리스트에서 글쓰기 버튼 클릭 시도
    if list_url and goto_write_from_list(drv, list_url, bid):
        log("✅ 글쓰기 페이지(리스트→버튼) 진입 성공")
        return

    # C) write.php 직접 오픈 (미로그인/권한 알럿이면 처리 후 재시도)
    safe_get(drv, write_url)
    if "board_write.php" not in drv.current_url:
        # 권한 문제 등으로 리다이렉트 되었을 수 있음 → 로그인 보장 후 재시도
        ensure_login(drv, list_url or "https://zae-da.com/bbs/list.php?boardid=" + (bid or ""), write_url)
        safe_get(drv, write_url)

    if "board_write.php" not in drv.current_url:
        raise RuntimeError("글쓰기 페이지로 진입하지 못했습니다.")


# ──────────────────────────────
# 입력/제출
# ──────────────────────────────
def fill_title(drv, title: str):
    inputs = [
        "input[name='wr_subject']",
        "input[name='subject']",
        "input[name='title']",
        "input[type='text']#wr_subject",
    ]
    ti, sel = find_first(drv, inputs, By.CSS_SELECTOR, wait_s=10)
    ti.clear()
    ti.send_keys(title)
    log("제목 입력 완료 ✓")


def fill_body(drv, body: str):
    """textarea → contenteditable → iframe 순으로 시도."""
    # 1) textarea
    try:
        ta, _ = find_first(
            drv,
            ["textarea[name='wr_content']", "textarea[name='content']", "textarea#wr_content", "textarea"],
            By.CSS_SELECTOR,
            wait_s=4,
        )
        ta.clear()
        ta.send_keys(body)
        log("본문 입력 완료 ✓ (textarea)")
        return
    except Exception:
        pass

    # 2) contenteditable
    try:
        ed, _ = find_first(drv, ["div[contenteditable='true']"], By.CSS_SELECTOR, wait_s=3)
        drv.execute_script("arguments[0].innerHTML = arguments[1];", ed, body.replace("\n", "<br>"))
        log("본문 입력 완료 ✓ (contenteditable)")
        return
    except Exception:
        pass

    # 3) iframe 에디터들 순회
    iframes = drv.find_elements(By.TAG_NAME, "iframe")
    for idx, ifr in enumerate(iframes):
        try:
            drv.switch_to.frame(ifr)
            # 에디터 내부 body/iframe 편집 영역 탐색
            try:
                editable = drv.find_elements(By.CSS_SELECTOR, "[contenteditable='true'], body")
                if editable:
                    el = editable[0]
                    # body의 경우 .innerHTML 세팅
                    drv.execute_script("arguments[0].innerHTML = arguments[1];", el, body.replace("\n", "<br>"))
                    log(f"본문 입력 완료 ✓ (iframe #{idx})")
                    drv.switch_to.default_content()
                    return
            finally:
                drv.switch_to.default_content()
        except Exception:
            drv.switch_to.default_content()
            continue

    raise NoSuchElementException("본문 입력 영역을 찾지 못했습니다. (textarea/contenteditable/iframe 불가)")


def submit_post(drv):
    # 등록/작성/저장 버튼
    sels = [
        "//button[contains(.,'등록') or contains(.,'작성') or contains(.,'저장')]",
        "//input[@type='submit']",
        "//a[contains(@onclick,'write') and (contains(.,'등록') or contains(.,'작성'))]",
    ]
    for xp in sels:
        try:
            btn = WebDriverWait(drv, 6).until(EC.element_to_be_clickable((By.XPATH, xp)))
            btn.click()
            time.sleep(0.8)
            accept_all_alerts(drv)
            log("등록 버튼 클릭 ✓")
            return
        except Exception:
            continue
    raise NoSuchElementException("등록/작성 버튼을 찾지 못했습니다.")


# ──────────────────────────────
# 크롬 드라이버
# ──────────────────────────────
def setup_driver() -> Chrome:
    load_dotenv()  # .env 읽기

    user_data_dir = os.getenv("CHROME_USER_DATA_DIR", "").strip()
    profile_dir   = os.getenv("CHROME_PROFILE", "").strip()
    fallback_dir  = os.getenv("CHROME_FALLBACK_DIR", "").strip()

    def _make_options(ud: str | None, prof: str | None) -> ChromeOptions:
        opts = ChromeOptions()
        if ud:
            opts.add_argument(f"--user-data-dir={ud}")
        if prof:
            # Windows의 멀티 프로필: "User Data" + "Profile xx"
            opts.add_argument(f"--profile-directory={prof}")

        # 안정화 옵션 (Windows)
        opts.add_argument("--start-maximized")
        opts.add_experimental_option("excludeSwitches", ["enable-automation", "enable-logging"])
        opts.add_experimental_option("useAutomationExtension", False)
        # 암시적 크래시 방지용(불필요한 경우도 있으나 무해)
        opts.add_argument("--disable-notifications")
        opts.add_argument("--disable-popup-blocking")

        return opts

    def _launch(opts: ChromeOptions) -> Chrome:
        return Chrome(service=Service(ChromeDriverManager().install()), options=opts)

    # 1차: 환경변수의 실제 프로필로 시도
    try:
        if user_data_dir:
            log("기존 브라우저 세션에서 여는 중입니다.")
        drv = _launch(_make_options(user_data_dir or None, profile_dir or None))
        return drv
    except (SessionNotCreatedException, WebDriverException) as e:
        msg = f"{e}"
        log(f"[chrome] primary profile failed → {msg}")

    # 2차: 폴백 프로필로 시도
    if not fallback_dir:
        # 폴백 경로 기본값
        fallback_dir = r"C:\ChromeProfiles\zaeda_selenium"
    Path(fallback_dir).mkdir(parents=True, exist_ok=True)
    drv = _launch(_make_options(fallback_dir, None))
    log(f"[chrome] fallback profile launched: {fallback_dir}\n  ↳ 폴백 창에서 재다몰에 1회 로그인해 두면 이후 자동 유지됩니다.")
    return drv


# ──────────────────────────────
# 메인
# ──────────────────────────────
def main():
    load_dotenv()
    ap = argparse.ArgumentParser()
    ap.add_argument("--url", required=True, help="글쓰기 폼 URL 예) https://zae-da.com/m/bbs/board_write.php?boardid=41")
    ap.add_argument("--list-url", default=None, help="게시판 리스트 URL 예) https://zae-da.com/bbs/list.php?boardid=41")
    args = ap.parse_args()

    # 엑셀에서 1건 꺼내오기
    wb, ws, row, title, body = load_next_row()
    if not row:
        log("대기 중인 업로드 건이 없습니다.")
        return

    # 크롬 구동
    drv = setup_driver()
    try:
        # 권한 알럿/미로그인 대비: 글쓰기 전 로그인 보장
        list_url = args.list_url or "https://zae-da.com"  # 최소 홈이라도 전달
        ensure_login(drv, list_url, args.url)

        # 글쓰기 페이지 진입 (리스트→버튼 우선, 실패 시 직접 진입)
        ensure_write_page(drv, args.list_url, args.url)

        # 제목/본문 입력
        fill_title(drv, title)
        fill_body(drv, body)

        # 제출
        submit_post(drv)

        # 완료 처리
        mark_done(wb, ws, row)
        log("✅ 업로드 완료 → DONE 처리")
    except UnexpectedAlertPresentException:
        # 권한/세션 알럿 등: 가능한 한 수락하고 종료
        try:
            accept_all_alerts(drv)
        except Exception:
            pass
        log("❌ 알럿으로 인해 제출이 중단되었습니다.")
        raise
    finally:
        try:
            # 닫지 않고 남겨두고 싶으면 주석 처리
            drv.quit()
        except Exception:
            pass


if __name__ == "__main__":
    main()
