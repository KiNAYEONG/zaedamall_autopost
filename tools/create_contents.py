# -*- coding: utf-8 -*-
"""
지침 기반 콘텐츠 생성기 (재다몰 자동 포스팅용)
- .env 에서 모델 키/옵션 불러오기
- docs/data.xlsx 에 비어있는 제목/본문을 지침에 맞춰 자동 채움
- 제목 규칙: `[카테고리1/카테고리2] 제목` (길이 22~30자, 금지어 필터)
- 본문 구조·톤&매너·디스클레이머 자동 반영
"""

from __future__ import annotations
import os, re, datetime, argparse
from pathlib import Path
from typing import Optional, Tuple
from dotenv import load_dotenv
import openpyxl

# ── 경로/상수 ─────────────────────────────
ROOT = Path(__file__).resolve().parent.parent
DOCS = ROOT / "docs"
DOCS.mkdir(exist_ok=True)
XLSX = DOCS / "data.xlsx"

TITLE_MIN = 22
TITLE_MAX = 30

CATEGORIES = {
    "만성질환 관리": ["당뇨 관리", "고혈압 관리", "비만 관리", "치매 관리"],
    "마음과 몸 관리": ["불면증", "만성피로", "소화불량", "두통", "우울"],
    "생활습관 관리": ["식습관", "운동 습관", "배변 습관", "음주 습관", "금연"],
}

FORBIDDEN_WORDS = ["100% 예방", "충격", "완치"]

# ── 모델 클라이언트 로드 ─────────────────
def _load_model():
    try:
        from tools.gemini_client import generate_text  # type: ignore
        return ("gemini", generate_text)
    except Exception:
        try:
            from gemini_client import generate_text  # type: ignore
            return ("gemini", generate_text)
        except Exception:
            pass

    # Fallback dummy
    def _fallback(prompt: str, **kwargs) -> str:
        return (
            "후크: 요즘 일상이 바쁜데도 증상 때문에 힘드시죠? "
            "오늘은 작은 습관이 큰 변화를 만드는 방법을 소개합니다.\n\n"
            "왜 중요한가: 몸의 균형과 생활습관이 건강 전반에 미치는 영향은 크며, "
            "연구에서도 생활습관 개선이 중요한 요인으로 보고됩니다.\n\n"
            "1) 물 마시기 루틴 만들기 💧 …\n"
            "2) 가벼운 걷기 습관 들이기 🚶 …\n"
            "3) 취침 전 휴대폰 줄이기 🌙 …\n\n"
            "주의사항: 약물 복용 중이거나 기존 질환이 있다면 전문가와 상담하세요.\n\n"
            "요약: 오늘부터 작은 실천으로도 건강 변화를 느낄 수 있습니다. 😊\n\n"
            "근거자료:\n- WHO 가이드\n- 질병관리청 자료\n\n"
            "이 글은 일반적인 건강 정보를 제공하기 위한 것이며, 의료적 진단이나 치료를 대신하지 않습니다. "
            "개인별 상태에 따라 전문가 상담이 필요할 수 있습니다."
        )
    return ("fallback", _fallback)

MODEL_NAME, MODEL_FN = _load_model()

# ── 프롬프트 ─────────────────────────────
GUIDELINE_PROMPT = """
당신은 블로그 포스트 작성기입니다. 한국어로 작성합니다.

[톤 & 매너]
- 대화체·공감형: 질문 → 공감 → 문제 정의
- 따뜻하고 격려: 공포·경고 금지
- 쉽게 풀기: 한글 우선, 영어는 보조
- 과학적이되 가볍게: 연구·기관 자료 1–2문장 인용
- 한국 맥락 반영: 밥·국·반찬·출퇴근·카톡 사례

[제목 규칙]
- 길이: 22–30자
- 1문장, 질문형/결과형/숫자형 중 택1
- 금지: 100% 예방, 충격, 완치
- 최종 제목은 반드시 [카테고리1/카테고리2] 접두 포함

[본문 구조·분량 ≈ 2000자]
1) 후크(250–300자)
2) 왜 중요한가(400–450자)
3) 핵심 행동 3가지(각 350–400자)
4) 주의사항(150–200자)
5) 요약(200–250자)
6) 근거자료(3–6개)
+ 디스클레이머 자동 추가

요청 주제: {topic}
카테고리1: {cat1}
카테고리2: {cat2}

출력 형식:
<제목>
<본문 전체>
"""

# ── 유틸 ─────────────────────────────
def ensure_workbook() -> openpyxl.Workbook:
    if XLSX.exists():
        return openpyxl.load_workbook(XLSX)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "posts"
    ws.append(["제목", "본문", "상태", "업데이트시각", "이미지검색어", "카테고리"])
    wb.save(XLSX)
    return wb

def sanitize_title(title: str) -> str:
    for bad in FORBIDDEN_WORDS:
        title = title.replace(bad, "")
    return re.sub(r"\s+", " ", title.strip())

def clip_title_len(title: str) -> str:
    if len(title) > TITLE_MAX:
        title = title[:TITLE_MAX]
    if len(title) < TITLE_MIN:
        title += " 시작해 보세요"
        if len(title) > TITLE_MAX:
            title = title[:TITLE_MAX]
    return title

def extract_title_and_body(raw: str) -> Tuple[str, str]:
    parts = raw.strip().split("\n", 1)
    return parts[0].strip(), parts[1].strip() if len(parts) > 1 else ""

def wrap_title_with_categories(title: str, cat1: str, cat2: str) -> str:
    prefix = f"[{cat1}/{cat2}] "
    if not title.startswith(prefix):
        title = prefix + title
    return clip_title_len(sanitize_title(title))

# ── 생성 로직 ─────────────────────────────
def generate_post(cat1: str, cat2: str, topic: Optional[str] = None) -> Tuple[str, str]:
    topic = topic or f"{cat2} 관리 가이드"
    prompt = GUIDELINE_PROMPT.format(topic=topic, cat1=cat1, cat2=cat2)

    text = MODEL_FN(prompt, max_output_tokens=2200) if MODEL_NAME == "gemini" else MODEL_FN(prompt)
    title, body = extract_title_and_body(text)

    title = wrap_title_with_categories(title, cat1, cat2)

    disclaimer = (
        "이 글은 일반적인 건강 정보를 제공하기 위한 것이며, 의료적 진단이나 치료를 대신하지 않습니다. "
        "개인별 상태에 따라 전문가 상담이 필요할 수 있습니다."
    )
    if disclaimer not in body:
        body = body.rstrip() + "\n\n" + disclaimer

    return title, body

# ── 메인 ─────────────────────────────
def main():
    load_dotenv()
    ap = argparse.ArgumentParser()
    ap.add_argument("--only-empty", action="store_true", help="빈 행만 채우기")
    ap.add_argument("--count", type=int, default=1, help="생성할 개수 (0=카테고리 전체 1회전)")
    ap.add_argument("--topic", type=str, default="", help="사용자 지정 주제")
    args = ap.parse_args()

    wb = ensure_workbook()
    ws = wb.active

    # 기존 제목 중복 방지
    existing_titles = {(ws[f"A{i}"].value or "").strip() for i in range(2, ws.max_row + 1)}

    generated = 0
    to_generate = []

    if args.count > 0:
        for cat1, subcats in CATEGORIES.items():
            for cat2 in subcats:
                to_generate.append((cat1, cat2))
                if len(to_generate) >= args.count:
                    break
            if len(to_generate) >= args.count:
                break
    else:
        for cat1, subcats in CATEGORIES.items():
            for cat2 in subcats:
                to_generate.append((cat1, cat2))

    for cat1, cat2 in to_generate:
        title, body = generate_post(cat1, cat2, args.topic or None)
        if title in existing_titles:
            print(f"⚠️ 중복 건너뜀: {title}")
            continue

        row = ws.max_row + 1
        ws[f"A{row}"] = title
        ws[f"B{row}"] = body
        ws[f"C{row}"] = "미발행"
        ws[f"D{row}"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
        ws[f"E{row}"] = cat2  # 이미지검색어
        ws[f"F{row}"] = f"{cat1}/{cat2}"
        generated += 1

    wb.save(XLSX)
    print(f"✅ 생성 완료: {generated}건 (모델: {MODEL_NAME}) → {XLSX}")

if __name__ == "__main__":
    main()
